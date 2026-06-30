"""Excel import helpers for the constructor library.

Supports two workbook contracts:
- legacy: Russian 3-sheet template used by the deprecated endpoint.
- developer: new 5-sheet workbook for candidate/developer profiles.
"""
from __future__ import annotations

import io
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TRUE_SET = {"да", "yes", "true", "1", "+", "✓", "истина"}

LEGACY_KIND_MAP = {
    "общая": "common",
    "общие": "common",
    "common": "common",
    "профессиональная": "professional",
    "проф": "professional",
    "professional": "professional",
}
LEGACY_QTYPE_MAP = {
    "один вариант": "single_choice",
    "одиночный": "single_choice",
    "single": "single_choice",
    "single_choice": "single_choice",
    "несколько вариантов": "multiple_choice",
    "мультивыбор": "multiple_choice",
    "multiple": "multiple_choice",
    "multiple_choice": "multiple_choice",
    "шкала": "scale",
    "scale": "scale",
    "открытый": "open",
    "open": "open",
}
LEGACY_TARGET_MAP = {
    "кандидат": "candidate",
    "кандидаты": "candidate",
    "candidate": "candidate",
    "сотрудник": "employee",
    "сотрудники": "employee",
    "employee": "employee",
    "группа": "group",
    "групповая": "group",
    "group": "group",
}

NEW_KIND_MAP = {"hard": "professional", "soft": "common"}
NEW_QTYPE_MAP = {
    "choice": "single_choice",
    "sjt": "single_choice",
    "case": "single_choice",
    "likert": "single_choice",
    "attention_check": "single_choice",
    "open": "open",
}


def _txt(v):
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip() or None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() or None


def _int(v, default=0):
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def _bool(v):
    value = _txt(v)
    return value is not None and value.lower() in TRUE_SET


def _rows(ws, headers):
    out = []
    it = ws.iter_rows(values_only=True)
    try:
        next(it)
    except StopIteration:
        return out
    for raw in it:
        if raw is None or not any(c is not None and str(c).strip() for c in raw):
            continue
        row = {headers[i]: (raw[i] if i < len(raw) else None) for i in range(len(headers))}
        out.append(row)
    return out


def _write_headers(ws, headers):
    header_fill = PatternFill("solid", fgColor="0A0F1E")
    header_font = Font(color="FFFFFF", bold=True)
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = 26
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _option_code_index(code: str | None) -> int:
    if not code:
        return 10_000
    code = code.strip().upper()
    if len(code) == 1 and "A" <= code <= "Z":
        return ord(code) - ord("A")
    return 10_000


# ---------------------------------------------------------------------------
# Legacy workbook

LEGACY_SHEETS = {
    "Компетенции": ["ID компетенции", "Название", "Тип (общая/профессиональная)", "Описание"],
    "Вопросы": [
        "ID компетенции",
        "ID вопроса",
        "Вопрос",
        "Тип (один вариант/несколько/шкала/открытый)",
        "Вариант ответа",
        "Баллы",
        "Верный (да/нет)",
        "Красный флаг (да/нет)",
        "Шкала мин",
        "Шкала макс",
        "Эталон ответа (AI)",
    ],
    "Профили": [
        "ID профиля",
        "Название профиля",
        "Категория",
        "Для кого (кандидат/сотрудник/группа)",
        "Длительность, мин",
        "ID компетенции",
        "Вес",
    ],
}


def build_legacy_template_bytes() -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    for name, headers in LEGACY_SHEETS.items():
        ws = wb.create_sheet(name)
        _write_headers(ws, headers)

    ws = wb["Компетенции"]
    ws.append(["C1", "Работа с возражениями", "профессиональная", "Умение отрабатывать отказы клиента"])
    ws.append(["C2", "Ответственность", "общая", "Доведение задач до результата"])

    ws = wb["Вопросы"]
    ws.append(["C1", "Q1", "Клиент говорит «дорого». Ваши действия?", "один вариант", "Выясню ценность и предложу варианты", 4, "да", "нет", "", "", ""])
    ws.append(["C1", "Q1", "", "", "Сразу дам максимальную скидку", 0, "нет", "да", "", "", ""])
    ws.append(["C1", "Q1", "", "", "Отпущу клиента", 0, "нет", "нет", "", "", ""])
    ws.append(["C2", "Q2", "Насколько вы согласны: «Я довожу задачи до конца»", "шкала", "", "", "", "", 1, 5, ""])
    ws.append(["C2", "Q3", "Опишите случай, когда вы взяли ответственность на себя", "открытый", "", "", "", "", "", "", "Конкретный пример с результатом и выводами"])

    ws = wb["Профили"]
    ws.append(["P1", "Менеджер по продажам", "Коммерция", "кандидат", 12, "C1", 1])
    ws.append(["P1", "Менеджер по продажам", "Коммерция", "кандидат", 12, "C2", 1])

    guide = wb.create_sheet("Инструкция")
    guide["A1"] = "Как заполнять"
    guide["A1"].font = Font(bold=True, size=13)
    tips = [
        "1. Лист «Компетенции»: по строке на компетенцию. ID - любой ваш уникальный код.",
        "2. Лист «Вопросы»: по строке на вариант ответа. Строки одного вопроса имеют один ID вопроса.",
        "3. Лист «Профили»: по строке на пару профиль+компетенция. ID компетенции ссылается на лист «Компетенции».",
        "4. Повторная загрузка файла безопасна: записи обновляются, а не дублируются.",
    ]
    for i, tip in enumerate(tips, start=3):
        guide.cell(row=i, column=1, value=tip)
    guide.column_dimensions["A"].width = 110

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_legacy_workbook(content: bytes) -> tuple[dict, list[str]]:
    errors: list[str] = []
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        return {}, [f"Не удалось открыть файл: {exc}. Нужен .xlsx по шаблону."]

    for required in ("Компетенции", "Вопросы", "Профили"):
        if required not in wb.sheetnames:
            errors.append(f"Нет листа «{required}». Используйте актуальный шаблон.")
    if errors:
        return {}, errors

    comp_rows = _rows(wb["Компетенции"], LEGACY_SHEETS["Компетенции"])
    q_rows = _rows(wb["Вопросы"], LEGACY_SHEETS["Вопросы"])
    p_rows = _rows(wb["Профили"], LEGACY_SHEETS["Профили"])

    competencies = []
    for row in comp_rows:
        cid = _txt(row["ID компетенции"])
        title = _txt(row["Название"])
        if not cid or not title:
            errors.append(f"Компетенция без ID или названия: {row}")
            continue
        competencies.append(
            {
                "ext_id": cid,
                "title": title,
                "kind": LEGACY_KIND_MAP.get((_txt(row["Тип (общая/профессиональная)"]) or "").lower(), "professional"),
                "description": _txt(row["Описание"]),
            }
        )

    questions: dict[str, dict] = {}
    for row in q_rows:
        qid = _txt(row["ID вопроса"])
        cid = _txt(row["ID компетенции"])
        if not qid or not cid:
            errors.append(f"Вопрос без ID вопроса/компетенции: {row}")
            continue
        q = questions.setdefault(
            qid,
            {
                "ext_id": qid,
                "competency_ext_id": cid,
                "text": None,
                "type": "single_choice",
                "scale_min": None,
                "scale_max": None,
                "ai_reference": None,
                "options": [],
            },
        )
        if _txt(row["Вопрос"]):
            q["text"] = _txt(row["Вопрос"])
        qtype = _txt(row["Тип (один вариант/несколько/шкала/открытый)"])
        if qtype:
            q["type"] = LEGACY_QTYPE_MAP.get(qtype.lower(), q["type"])
        if row["Шкала мин"] not in (None, ""):
            q["scale_min"] = _int(row["Шкала мин"], 1)
        if row["Шкала макс"] not in (None, ""):
            q["scale_max"] = _int(row["Шкала макс"], 5)
        if _txt(row["Эталон ответа (AI)"]):
            q["ai_reference"] = _txt(row["Эталон ответа (AI)"])
        opt_text = _txt(row["Вариант ответа"])
        if opt_text:
            q["options"].append(
                {
                    "text": opt_text,
                    "score": _int(row["Баллы"], 0),
                    "is_correct": _bool(row["Верный (да/нет)"]),
                    "is_red_flag": _bool(row["Красный флаг (да/нет)"]),
                }
            )

    profiles: dict[str, dict] = {}
    for row in p_rows:
        pid = _txt(row["ID профиля"])
        title = _txt(row["Название профиля"])
        if not pid or not title:
            errors.append(f"Профиль без ID/названия: {row}")
            continue
        p = profiles.setdefault(
            pid,
            {
                "ext_id": pid,
                "title": title,
                "category": _txt(row["Категория"]),
                "target_type": LEGACY_TARGET_MAP.get((_txt(row["Для кого (кандидат/сотрудник/группа)"]) or "").lower(), "candidate"),
                "duration": _int(row["Длительность, мин"], 0),
                "competencies": [],
            },
        )
        cid = _txt(row["ID компетенции"])
        if cid:
            p["competencies"].append({"competency_ext_id": cid, "weight": _int(row["Вес"], 1) or 1})

    for q in questions.values():
        if not q["text"]:
            errors.append(f"Вопрос {q['ext_id']} без текста — пропущен.")
    questions = {k: v for k, v in questions.items() if v["text"]}

    return {"competencies": competencies, "questions": list(questions.values()), "profiles": list(profiles.values())}, errors


# Backward-compatible aliases for older call sites.
build_template_bytes_legacy = build_legacy_template_bytes
parse_workbook_legacy = parse_legacy_workbook


# ---------------------------------------------------------------------------
# New workbook

NEW_SHEETS = {
    "profiles": [
        "profile_id",
        "name",
        "category",
        "cat_code",
        "level",
        "total_questions",
        "time_min",
        "threshold_recommended_pct",
        "threshold_conditional_pct",
        "hard_weight_pct",
        "soft_weight_pct",
        "mandatory_conditions",
        "variations",
        "version",
    ],
    "competencies": ["profile_id", "comp_code", "comp_name", "comp_type", "weight_pct", "n_questions", "min_score_pct", "description"],
    "questions": [
        "profile_id",
        "quest_id",
        "comp_code",
        "seq_in_comp",
        "q_type",
        "answer_format",
        "difficulty",
        "level_tag",
        "text",
        "option_a",
        "option_b",
        "option_c",
        "option_d",
        "score_a",
        "score_b",
        "score_c",
        "score_d",
        "comment",
    ],
    "profile_comp_questions": ["profile_id", "comp_code", "comp_type", "comp_weight_pct", "quest_id", "q_type", "answer_format", "difficulty", "level_tag"],
    "answer_options": ["profile_id", "quest_id", "option_code", "option_text", "score"],
}


def build_candidates_developer_template_bytes() -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    for name, headers in NEW_SHEETS.items():
        ws = wb.create_sheet(name)
        _write_headers(ws, headers)

    # Заполненный пример одного профиля, чтобы шаблон был самодостаточным.
    wb["profiles"].append([
        "CAND-IT-04-001", "Backend-разработчик (Python)", "IT-специалисты", "C04", "L2–L3",
        2, 60, 80, 60, 85, 15, "C1(Python core)≥60%", "—", "1.0",
    ])
    wb["competencies"].append([
        "CAND-IT-04-001", "C1", "Python core", "hard", 100, 2, 60,
        "Синтаксис, типы, ООП, декораторы, генераторы, память",
    ])
    wb["questions"].append([
        "CAND-IT-04-001", "C1-01", "C1", 1, "choice", "Один вариант ответа", "Базовый", "all",
        "Какой из типов данных в Python является неизменяемым (immutable)?",
        "list", "dict", "tuple", "set", 0, 0, 5, 0,
        "tuple — неизменяемый; list, dict, set — изменяемые.",
    ])
    wb["questions"].append([
        "CAND-IT-04-001", "C1-02", "C1", 2, "choice", "Один вариант ответа", "Средний", "all",
        "Что выведет код: x=[1,2,3]; y=x; y.append(4); print(x)?",
        "[1, 2, 3]", "[1, 2, 3, 4]", "[4]", "Ошибка", 0, 5, 0, 0,
        "Присваивание списка создаёт ссылку, не копию: x и y — один объект.",
    ])
    for pid, qid, ct, cw, q_type in (
        ("CAND-IT-04-001", "C1-01", "hard", 100, "choice"),
        ("CAND-IT-04-001", "C1-02", "hard", 100, "choice"),
    ):
        wb["profile_comp_questions"].append([pid, "C1", ct, cw, qid, q_type, "Один вариант ответа", "Базовый", "all"])
    for qid, opts in (
        ("C1-01", [("A", "list", 0), ("B", "dict", 0), ("C", "tuple", 5), ("D", "set", 0)]),
        ("C1-02", [("A", "[1, 2, 3]", 0), ("B", "[1, 2, 3, 4]", 5), ("C", "[4]", 0), ("D", "Ошибка", 0)]),
    ):
        for code, text, score in opts:
            wb["answer_options"].append(["CAND-IT-04-001", qid, code, text, score])

    guide = wb.create_sheet("instructions")
    guide["A1"] = "Шаблон импорта профилей (формат candidates/developer)"
    guide["A1"].font = Font(bold=True, size=13)
    notes = [
        "1. Точные имена листов: profiles, competencies, questions, profile_comp_questions, answer_options.",
        "2. profile_id — стабильный ключ профиля. comp_code и quest_id уникальны только внутри одного профиля.",
        "3. competencies и questions профиль-специфичны: один comp_code может встречаться в разных профилях с разным содержанием.",
        "4. profile_comp_questions задаёт порядок вопросов внутри компетенции.",
        "5. answer_options — основной источник вариантов ответа; колонки option_a..d в questions — компактный запасной вариант.",
        "6. comp_type: hard (профессиональная) или soft (общая). q_type: choice / sjt / case / likert / attention_check / open.",
        "7. Повторная загрузка обновляет записи по ключам, а не дублирует. Строки-примеры можно заменить своими данными.",
    ]
    for i, note in enumerate(notes, start=3):
        guide.cell(row=i, column=1, value=note)
    guide.column_dimensions["A"].width = 115

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _normalize_new_qtype(value: str | None) -> str:
    key = (_txt(value) or "").lower()
    return NEW_QTYPE_MAP.get(key, "single_choice")


def _profile_summary(row: dict) -> str | None:
    parts = []
    level = _txt(row.get("level"))
    time_min = _int(row.get("time_min"), 0)
    rec = _int(row.get("threshold_recommended_pct"), 0)
    cond = _int(row.get("threshold_conditional_pct"), 0)
    hard = _int(row.get("hard_weight_pct"), 0)
    soft = _int(row.get("soft_weight_pct"), 0)
    if level:
        parts.append(level)
    if time_min:
        parts.append(f"{time_min} мин")
    if rec:
        parts.append(f"рекоменд. ≥ {rec}%")
    if cond:
        parts.append(f"условно ≥ {cond}%")
    if hard or soft:
        parts.append(f"hard/soft {hard}/{soft}%")
    mandatory = _txt(row.get("mandatory_conditions"))
    if mandatory and mandatory != "—":
        parts.append(mandatory)
    return " · ".join(parts) if parts else None


def parse_workbook(content: bytes) -> tuple[dict, list[str]]:
    errors: list[str] = []
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        return {}, [f"Не удалось открыть файл: {exc}. Нужен .xlsx по новому шаблону."]

    required = ("profiles", "competencies", "questions", "profile_comp_questions", "answer_options")
    for sheet in required:
        if sheet not in wb.sheetnames:
            errors.append(f"Нет листа «{sheet}». Используйте актуальный шаблон.")
    if errors:
        return {}, errors

    profile_rows = _rows(wb["profiles"], NEW_SHEETS["profiles"])
    comp_rows = _rows(wb["competencies"], NEW_SHEETS["competencies"])
    question_rows = _rows(wb["questions"], NEW_SHEETS["questions"])
    pcq_rows = _rows(wb["profile_comp_questions"], NEW_SHEETS["profile_comp_questions"])
    option_rows = _rows(wb["answer_options"], NEW_SHEETS["answer_options"])

    profiles_meta: dict[str, dict] = {}
    profile_order: list[str] = []
    for row in profile_rows:
        pid = _txt(row["profile_id"])
        name = _txt(row["name"])
        if not pid or not name:
            errors.append(f"Профиль без profile_id или name: {row}")
            continue
        profiles_meta[pid] = {
            "ext_id": pid,
            "title": name,
            "category": _txt(row["category"]),
            "target_type": "candidate",
            "duration": _int(row["time_min"], 0),
            "summary": _profile_summary(row),
            "competencies": [],
        }
        profile_order.append(pid)

    comp_meta: dict[tuple[str, str], dict] = {}
    comp_order: dict[str, list[str]] = defaultdict(list)
    for row in comp_rows:
        pid = _txt(row["profile_id"])
        code = _txt(row["comp_code"])
        name = _txt(row["comp_name"])
        if not pid or not code or not name:
            errors.append(f"Компетенция без profile_id/comp_code/comp_name: {row}")
            continue
        comp_meta[(pid, code)] = {
            "ext_id": f"{pid}:{code}",
            "title": name,
            "kind": NEW_KIND_MAP.get((_txt(row["comp_type"]) or "").lower(), "professional"),
            "weight": _int(row["weight_pct"], 0),
            "n_questions": _int(row["n_questions"], 0),
            "min_score_pct": _int(row["min_score_pct"], 0),
            "description": _txt(row["description"]),
            "code": code,
            "questions": [],
        }
        if code not in comp_order[pid]:
            comp_order[pid].append(code)

    questions_meta: dict[tuple[str, str], dict] = {}
    for row in question_rows:
        pid = _txt(row["profile_id"])
        qid = _txt(row["quest_id"])
        code = _txt(row["comp_code"])
        text = _txt(row["text"])
        if not pid or not qid or not code or not text:
            errors.append(f"Вопрос без profile_id/quest_id/comp_code/text: {row}")
            continue
        raw_q_type = _txt(row["q_type"])
        qtype = _normalize_new_qtype(raw_q_type)
        options = []
        for opt_code, opt_text, score_key in (
            ("A", row["option_a"], row["score_a"]),
            ("B", row["option_b"], row["score_b"]),
            ("C", row["option_c"], row["score_c"]),
            ("D", row["option_d"], row["score_d"]),
        ):
            opt_text = _txt(opt_text)
            if opt_text is None:
                continue
            options.append(
                {
                    "text": opt_text,
                    "score": _int(score_key, 0),
                    "is_correct": False,
                    "is_red_flag": False,
                    "option_code": opt_code,
                }
            )

        max_score = 5 if qtype == "open" else max((opt["score"] for opt in options), default=0)
        if qtype == "single_choice" and raw_q_type != "attention_check" and options:
            best_idx = max(range(len(options)), key=lambda i: (options[i]["score"], -i))
            for idx, opt in enumerate(options):
                opt["is_correct"] = idx == best_idx and options[best_idx]["score"] > 0
        if raw_q_type == "attention_check":
            best_idx = max(range(len(options)), key=lambda i: (options[i]["score"], -i)) if options else None
            for idx, opt in enumerate(options):
                opt["is_correct"] = idx == best_idx
                opt["is_red_flag"] = idx != best_idx
                opt["score"] = 0
            max_score = 0
        if qtype == "open":
            options = []

        questions_meta[(pid, qid)] = {
            "ext_id": f"{pid}:{qid}",
            "profile_id": pid,
            "quest_id": qid,
            "competency_code": code,
            "text": text,
            "type": qtype,
            "raw_q_type": raw_q_type,
            "answer_format": _txt(row["answer_format"]),
            "difficulty": _txt(row["difficulty"]),
            "level_tag": _txt(row["level_tag"]),
            "options": sorted(options, key=lambda opt: _option_code_index(opt.get("option_code"))),
            "scale_min": None,
            "scale_max": None,
            "ai_reference": _txt(row["comment"]),
            "ai_criteria": _txt(row["comment"]),
            "max_score": max_score,
            "seq_in_comp": _int(row["seq_in_comp"], 0),
        }

    options_by_question: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in option_rows:
        pid = _txt(row["profile_id"])
        qid = _txt(row["quest_id"])
        text = _txt(row["option_text"])
        if not pid or not qid or not text:
            errors.append(f"Вариант ответа без profile_id/quest_id/text: {row}")
            continue
        options_by_question[(pid, qid)].append(
            {
                "option_code": _txt(row["option_code"]),
                "text": text,
                "score": _int(row["score"], 0),
            }
        )

    pcq_by_comp: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in pcq_rows:
        pid = _txt(row["profile_id"])
        code = _txt(row["comp_code"])
        qid = _txt(row["quest_id"])
        if not pid or not code or not qid:
            errors.append(f"Связка profile/competency/question неполная: {row}")
            continue
        pcq_by_comp[(pid, code)].append(
            {
                "quest_id": qid,
                "comp_type": _txt(row["comp_type"]),
                "comp_weight_pct": _int(row["comp_weight_pct"], 0),
                "q_type": _normalize_new_qtype(row["q_type"]),
                "answer_format": _txt(row["answer_format"]),
                "difficulty": _txt(row["difficulty"]),
                "level_tag": _txt(row["level_tag"]),
            }
        )

    # Merge answer options from the dedicated sheet.
    for (pid, qid), q in questions_meta.items():
        opts = options_by_question.get((pid, qid))
        if not opts:
            continue
        opts = sorted(opts, key=lambda opt: _option_code_index(opt.get("option_code")))
        q["options"] = [{"text": opt["text"], "score": opt["score"], "is_correct": False, "is_red_flag": False} for opt in opts]
        if q["raw_q_type"] == "attention_check":
            best_idx = max(range(len(q["options"])), key=lambda i: (q["options"][i]["score"], -i))
            for idx, opt in enumerate(q["options"]):
                opt["is_red_flag"] = idx != best_idx
                opt["score"] = 0
            q["max_score"] = 0
        elif q["type"] == "single_choice":
            max_score = max((opt["score"] for opt in q["options"]), default=0)
            q["max_score"] = max_score
            if q["options"] and max_score > 0:
                best_idx = max(range(len(q["options"])), key=lambda i: (q["options"][i]["score"], -i))
                for idx, opt in enumerate(q["options"]):
                    opt["is_correct"] = idx == best_idx

    # Attach ordered questions to competencies and competencies to profiles.
    for pid in profile_order:
        profile = profiles_meta.get(pid)
        if profile is None:
            continue
        for comp_code in comp_order.get(pid, []):
            comp = comp_meta.get((pid, comp_code))
            if comp is None:
                continue
            rows = sorted(
                pcq_by_comp.get((pid, comp_code), []),
                key=lambda item: (
                    questions_meta.get((pid, item["quest_id"]), {}).get("seq_in_comp", 0),
                    item["quest_id"],
                ),
            )
            ordered_questions = []
            for link in rows:
                q = questions_meta.get((pid, link["quest_id"]))
                if q is None:
                    errors.append(f"Профиль {pid}, компетенция {comp_code}: вопрос {link['quest_id']} не найден в листе questions")
                    continue
                q = dict(q)
                q["type"] = link["q_type"] or q["type"]
                q["answer_format"] = link["answer_format"] or q["answer_format"]
                q["difficulty"] = link["difficulty"] or q["difficulty"]
                q["level_tag"] = link["level_tag"] or q["level_tag"]
                ordered_questions.append(q)
            comp = dict(comp)
            comp["questions"] = ordered_questions
            profile["competencies"].append(comp)

    for pid, profile in profiles_meta.items():
        if not profile["competencies"]:
            errors.append(f"Профиль {pid} не содержит компетенций.")
        for comp in profile["competencies"]:
            if not comp["questions"]:
                errors.append(f"Профиль {pid}, компетенция {comp['code']} не содержит вопросов.")
            if comp["weight"] <= 0:
                comp["weight"] = 1

    return {"profiles": list(profiles_meta.values())}, errors


build_candidates_developer_template = build_candidates_developer_template_bytes
parse_workbook_new = parse_workbook
