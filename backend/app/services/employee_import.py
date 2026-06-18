"""Импорт сотрудников из Excel: единый источник правды по колонкам шаблона.

Здесь же генерируется сам шаблон (GET /employees/import/template) и парсится
загруженный файл (POST /employees/import) — чтобы колонки шаблона и парсера
всегда совпадали.
"""
from __future__ import annotations

import io
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# (заголовок в шаблоне, поле, обязательное, подсказка)
COLUMNS: list[tuple[str, str, bool, str]] = [
    ("ФИО", "full_name", True, "Фамилия Имя Отчество полностью. Обязательное поле."),
    ("Email", "email", False, "Нужен для отправки оценок и отчётов."),
    ("Телефон", "phone", False, "Любой удобный формат."),
    ("Должность", "position", False, "Фактическая должность сотрудника."),
    ("Отдел", "department_name", False, "Отдел/команда. Создаётся автоматически, если новый."),
    ("Руководитель", "manager_name", False, "ФИО руководителя (как в колонке ФИО)."),
    ("Проект", "project", False, "Проект, направление или филиал."),
    ("Дата выхода", "start_date", False, "ГГГГ-ММ-ДД или ДД.ММ.ГГГГ. Запускает цикл адаптации."),
    ("Тип занятости", "employment_type", False, "Например: штат, ГПХ, стажёр."),
    ("Город", "city", False, "Город работы сотрудника."),
]

SHEET_NAME = "Сотрудники"

# Синонимы заголовков (нормализованные: lower + trim) → поле.
_HEADER_ALIASES: dict[str, str] = {}
for _title, _field, _req, _hint in COLUMNS:
    _HEADER_ALIASES[_title.strip().lower()] = _field
_HEADER_ALIASES.update({
    "ф.и.о.": "full_name", "фио сотрудника": "full_name", "сотрудник": "full_name",
    "полное имя": "full_name",
    "e-mail": "email", "почта": "email", "электронная почта": "email",
    "тел": "phone", "тел.": "phone", "phone": "phone",
    "позиция": "position",
    "департамент": "department_name", "подразделение": "department_name", "команда": "department_name",
    "менеджер": "manager_name", "manager": "manager_name",
    "направление": "project",
    "дата приёма": "start_date", "дата приема": "start_date", "дата найма": "start_date",
    "дата выхода на работу": "start_date", "start date": "start_date",
    "занятость": "employment_type", "тип занятости ": "employment_type",
    "city": "city",
})

_DATE_FORMATS = ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"]

# Поле-дата (парсится отдельно); остальные ячейки приводим к тексту.
_DATE_FIELD = "start_date"


def _to_text(value) -> str | None:
    """Приводит ячейку к строке. Excel часто отдаёт телефоны/числа как float —
    без этого Pydantic ругается «Input should be a valid string»."""
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() or None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip() or None


def _parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"не распознана дата «{text}» (нужен формат ГГГГ-ММ-ДД или ДД.ММ.ГГГГ)")


def build_template_bytes() -> bytes:
    """Генерирует .xlsx-шаблон с листом данных и листом-инструкцией."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    header_fill = PatternFill("solid", fgColor="0A0F1E")
    header_font = Font(color="FFFFFF", bold=True)
    for idx, (title, *_rest) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = 22
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    examples = [
        ["Иванов Иван Иванович", "ivanov@example.ru", "+7 900 000-00-00", "Менеджер по продажам",
         "Отдел продаж", "Петров Пётр Петрович", "B2B", "2026-06-01", "штат", "Москва"],
        ["Соколова Анна Сергеевна", "sokolova@example.ru", "+7 901 000-00-00", "HR-рекрутер",
         "HR", "", "Подбор", "2026-05-15", "штат", "Санкт-Петербург"],
    ]
    for r, row in enumerate(examples, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    guide = wb.create_sheet("Инструкция")
    guide.cell(row=1, column=1, value="Колонка").font = Font(bold=True)
    guide.cell(row=1, column=2, value="Что заполнить").font = Font(bold=True)
    guide.cell(row=1, column=3, value="Обязательное").font = Font(bold=True)
    for i, (title, _field, required, hint) in enumerate(COLUMNS, start=2):
        guide.cell(row=i, column=1, value=title)
        guide.cell(row=i, column=2, value=hint)
        guide.cell(row=i, column=3, value="да" if required else "нет")
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 60
    guide.column_dimensions["C"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_workbook(content: bytes) -> tuple[list[dict], list[str]]:
    """Парсит загруженный .xlsx → (строки, ошибки парсинга).

    Каждая строка — dict полей + служебное `_row` (номер строки в файле).
    Полностью пустые строки игнорируются.
    """
    errors: list[str] = []
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        return [], [f"Не удалось открыть файл: {exc}. Нужен .xlsx по шаблону."]

    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], ["Файл пустой — нет строки заголовков."]

    # Сопоставляем индекс колонки → поле по заголовку.
    col_to_field: dict[int, str] = {}
    for i, raw in enumerate(header or []):
        if raw is None:
            continue
        key = str(raw).strip().lower()
        field = _HEADER_ALIASES.get(key)
        if field:
            col_to_field[i] = field
    if "full_name" not in col_to_field.values():
        return [], ["В шаблоне не найдена колонка «ФИО». Используйте актуальный шаблон."]

    rows: list[dict] = []
    for r_idx, raw_row in enumerate(rows_iter, start=2):
        if raw_row is None:
            continue
        record: dict = {}
        for col_idx, field in col_to_field.items():
            val = raw_row[col_idx] if col_idx < len(raw_row) else None
            # Дату обрабатываем отдельно ниже; остальное приводим к тексту.
            record[field] = val if field == _DATE_FIELD else _to_text(val)
        # Полностью пустая строка — пропускаем молча.
        if not any(v is not None for v in record.values()):
            continue
        try:
            record[_DATE_FIELD] = _parse_date(record.get(_DATE_FIELD))
        except ValueError as exc:
            errors.append(f"Строка {r_idx}: {exc}")
            continue
        record["_row"] = r_idx
        rows.append(record)

    return rows, errors
