"""Импорт кандидатов из Excel: единый источник правды по колонкам шаблона.

Здесь генерируется шаблон (GET /candidates/import/template) и парсится
загруженный файл (POST /candidates/import) — колонки шаблона и парсера всегда
совпадают.
"""
from __future__ import annotations

import io
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# (заголовок в шаблоне, поле, обязательное, подсказка)
COLUMNS: list[tuple[str, str, bool, str]] = [
    ("ФИО", "full_name", True, "Фамилия Имя Отчество полностью. Обязательное поле."),
    ("Email", "email", False, "Нужен для отправки оценки и отчёта."),
    ("Телефон", "phone", False, "Любой удобный формат."),
    ("Вакансия", "vacancy_title", False, "Название вакансии. Создаётся автоматически, если новая."),
    ("Источник", "source", False, "Например: HeadHunter, SuperJob, Telegram, рекомендация."),
    ("Тип подбора", "selection_type", False, "Массовый или точечный подбор."),
    ("Дата отклика", "applied_at", False, "ГГГГ-ММ-ДД или ДД.ММ.ГГГГ."),
    ("Город", "city", False, "Город кандидата."),
    ("Комментарий", "notes", False, "Внутренняя заметка HR."),
]

SHEET_NAME = "Кандидаты"

_HEADER_ALIASES: dict[str, str] = {}
for _title, _field, _req, _hint in COLUMNS:
    _HEADER_ALIASES[_title.strip().lower()] = _field
_HEADER_ALIASES.update({
    "ф.и.о.": "full_name", "фио кандидата": "full_name", "кандидат": "full_name",
    "полное имя": "full_name",
    "e-mail": "email", "почта": "email", "электронная почта": "email",
    "тел": "phone", "тел.": "phone", "phone": "phone",
    "должность": "vacancy_title", "позиция": "vacancy_title", "вакансия/должность": "vacancy_title",
    "source": "source", "канал": "source",
    "тип подбора ": "selection_type", "подбор": "selection_type",
    "дата": "applied_at", "дата заявки": "applied_at", "дата отклика ": "applied_at",
    "city": "city",
    "заметка": "notes", "комментарий hr": "notes", "примечание": "notes",
})

_DATE_FORMATS = ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"]

# Поле-дата (парсится отдельно); остальные ячейки приводим к тексту.
_DATE_FIELD = "applied_at"


def _to_text(value) -> str | None:
    """Приводит ячейку к строке. Excel часто отдаёт телефоны/числа как float —
    без этого Pydantic ругается «Input should be a valid string»."""
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() or None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip() or None


def _parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"не распознана дата «{text}» (нужен формат ГГГГ-ММ-ДД или ДД.ММ.ГГГГ)")


def build_template_bytes() -> bytes:
    """Генерирует .xlsx-шаблон импорта кандидатов (данные + инструкция)."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    header_fill = PatternFill("solid", fgColor="0A0F1E")
    header_font = Font(color="FFFFFF", bold=True)
    for idx, (title, *_rest) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = 22
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    examples = [
        ["Кузнецов Артём Игоревич", "kuznetsov@example.ru", "+7 900 000-00-00",
         "Менеджер по продажам", "HeadHunter", "точечный", "2026-06-05", "Москва", "Сильное резюме"],
        ["Лебедева Ольга Павловна", "lebedeva@example.ru", "+7 901 000-00-00",
         "Оператор call-центра", "SuperJob", "массовый", "2026-06-08", "Казань", ""],
    ]
    for r, row in enumerate(examples, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    guide = wb.create_sheet("Инструкция")
    guide.cell(row=1, column=1, value="Колонка").font = Font(bold=True)
    guide.cell(row=1, column=2, value="Что заполнить").font = Font(bold=True)
    guide.cell(row=1, column=3, value="Обязательное").font = Font(bold=True)
    for i, (title, _field, required, hint) in enumerate(COLUMNS, start=2):
        guide.cell(row=i, column=1, value=title)
        guide.cell(row=i, column=2, value=hint)
        guide.cell(row=i, column=3, value="да" if required else "нет")
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 60
    guide.column_dimensions["C"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_workbook(content: bytes) -> tuple[list[dict], list[str]]:
    """Парсит .xlsx → (строки, ошибки парсинга). Пустые строки игнорируются."""
    errors: list[str] = []
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        return [], [f"Не удалось открыть файл: {exc}. Нужен .xlsx по шаблону."]

    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], ["Файл пустой — нет строки заголовков."]

    col_to_field: dict[int, str] = {}
    for i, raw in enumerate(header or []):
        if raw is None:
            continue
        field = _HEADER_ALIASES.get(str(raw).strip().lower())
        if field:
            col_to_field[i] = field
    if "full_name" not in col_to_field.values():
        return [], ["В шаблоне не найдена колонка «ФИО». Используйте актуальный шаблон."]

    rows: list[dict] = []
    for r_idx, raw_row in enumerate(rows_iter, start=2):
        if raw_row is None:
            continue
        record: dict = {}
        for col_idx, field in col_to_field.items():
            val = raw_row[col_idx] if col_idx < len(raw_row) else None
            record[field] = val if field == _DATE_FIELD else _to_text(val)
        if not any(v is not None for v in record.values()):
            continue
        try:
            record[_DATE_FIELD] = _parse_date(record.get(_DATE_FIELD))
        except ValueError as exc:
            errors.append(f"Строка {r_idx}: {exc}")
            continue
        record["_row"] = r_idx
        rows.append(record)

    return rows, errors
